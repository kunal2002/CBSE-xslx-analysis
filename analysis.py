from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('Result.xlsx')
ws = wb.active
wb1 = Workbook()
ws1 = wb1.active
ws.title = "Kunal's Analysis of Result"
headings = [301,28,39,65,37,48,64,41,74,42,43,44,83,54,55,28,30]
ws1.append(headings)
ctr = 2
# Nested for loop to find out all unique subject codes using set data structure
thisset = []
for row in range(6,480,2):
    for col in range(5,11):
        letter = get_column_letter(col)
        thisset.append(ws[letter + str(row)].value)
print(set(thisset))

#Identifying subject codes and putting marks in the column of the subject code, student wise
for row in range(6,479,2):
    for col in range(5,11):
        letter = get_column_letter(col)
        #print(ws[letter + str(row)].value,end = " ")
        if ws[letter + str(row)].value == 301:
            ws1['A' + str(ctr)] = ws[letter + str(row+1)].value
        elif ws[letter + str(row)].value == 28:
            ws1['B' + str(ctr)] = ws[letter + str(row + 1)].value
        elif ws[letter + str(row)].value == 39:
            ws1['C' + str(ctr)] = ws[letter + str(row + 1)].value
        elif ws[letter + str(row)].value == 65:
            ws1['D' + str(ctr)] = ws[letter + str(row + 1)].value
        elif ws[letter + str(row)].value == 37:
            ws1['E' + str(ctr)] = ws[letter + str(row + 1)].value
        elif ws[letter + str(row)].value == 48:
            ws1['F' + str(ctr)] = ws[letter + str(row + 1)].value
        elif ws[letter + str(row)].value == 64:
            ws1['G' + str(ctr)] = ws[letter + str(row + 1)].value
        elif ws[letter + str(row)].value == 41:
            ws1['H' + str(ctr)] = ws[letter + str(row + 1)].value
        elif ws[letter + str(row)].value == 74:
            ws1['I' + str(ctr)] = ws[letter + str(row + 1)].value
        elif ws[letter + str(row)].value == 42:
            ws1['J' + str(ctr)] = ws[letter + str(row + 1)].value
        elif ws[letter + str(row)].value == 43:
            ws1['K' + str(ctr)] = ws[letter + str(row + 1)].value
        elif ws[letter + str(row)].value == 44:
            ws1['L' + str(ctr)] = ws[letter + str(row + 1)].value
        elif ws[letter + str(row)].value == 83:
            ws1['M' + str(ctr)] = ws[letter + str(row + 1)].value
        elif ws[letter + str(row)].value == 54:
            ws1['N' + str(ctr)] = ws[letter + str(row + 1)].value
        elif ws[letter + str(row)].value == 55:
            ws1['O' + str(ctr)] = ws[letter + str(row + 1)].value
        elif ws[letter + str(row)].value == 28:
            ws1['P' + str(ctr)] = ws[letter + str(row + 1)].value
        elif ws[letter + str(row)].value == 30:
            ws1['Q' + str(ctr)] = ws[letter + str(row + 1)].value
    ctr = ctr + 1
wb1.save("KnResult1.xlsx")
wb1 = load_workbook("KnResult.xlsx")
ws1 = wb1.active
ws1.insert_cols(1)
ws1['A1'] = "Name"
ctr = 2
#print(ws['D6'].value)
for row in range(6,479,2):
    ws1['A' + str(ctr)] = ws['D' + str(row)].value
    ctr = ctr + 1
ws1.insert_cols(1)
ws1['A1'] = "Admn No."
ctr = 2
for row in range(6,479,2):
    ws1['A' + str(ctr)] = ws['B' + str(row)].value
    ctr = ctr + 1

wb1.save("KnResult.xlsx")
print(ws['I6'].value)
